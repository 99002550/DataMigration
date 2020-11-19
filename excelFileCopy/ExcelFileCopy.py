# importing openpyxl module
import openpyxl as xl
import xlsxwriter
import os

# opening the source excel file
filename = "C:\\Users\\99002631\\Documents\\psno_99002631.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# opening the destination excel file
filename1 = "C:\\Users\\99002631\\Desktop\\Sample.xlsx"

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file

# file exist with read and write permissions
if os.path.isfile(filename1) and os.access(filename1, os.R_OK):
    print('file exist\n')
    wb2 = xl.load_workbook(filename1)
    ws2 = wb2.active
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row=i, column=j)

            # writing the read value to destination excel file
            ws2.cell(row=i, column=j).value = c.value
    # saving the destination excel file
    wb2.save(str(filename1))

# file exist with but read and write permissions are not granted
elif os.path.isfile(filename1) and not os.access(filename1, os.R_OK):
    print('file access denied. operation ended\n')

# if file does not exist
elif not os.path.isfile(filename1):
    print('file created\n')
    # Creating the destination excel file
    workbook = xlsxwriter.Workbook('C:\\Users\\99002631\\Desktop\\Sample.xlsx')
    worksheet = workbook.add_worksheet()
    workbook.close()
    wb2 = xl.load_workbook(filename1)
    ws2 = wb2.active
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row=i, column=j)

            # writing the read value to destination excel file
            ws2.cell(row=i, column=j).value = c.value
    # saving the destination excel file
    wb2.save(str(filename1))

else:
    # if some error comes up
    print('Something went wrong\n')
